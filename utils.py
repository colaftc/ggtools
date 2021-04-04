from typing import List

from werkzeug.datastructures import FileStorage
from openpyxl.worksheet.worksheet import Worksheet
from collections import namedtuple
from openpyxl import load_workbook
import json
import os
import datetime

# tencent cloud sdk
from tencentcloud.common import credential
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.sms.v20190711 import sms_client, models
from models import Seller, Following, FollowStatusChoices, db


SentClient = namedtuple('SentClient', ['name', 'tel', 'address', 'industry', 'company'])

# const index for the file format
NAME_CELL_INDEX = 0
MOBILE_CELL_INDEX = 3
COMPANY_CELL_INDEX = 4
INDUSTRY_CELL_INDEX = 5
ADDRESS_CELL_INDEX = 22
DEFAULT_SHEET_NAME = 'Sheet0'


def init_credential(appid, secret):
    _cred = None

    def inner():
        nonlocal _cred
        if _cred is None:
            _cred = credential.Credential(appid, secret)
        return _cred
    return inner


def parse_client_list(
        req,
        file: FileStorage,
        add_86prefix: bool = True,
        name_cell_index=NAME_CELL_INDEX,
        mobile_cell_index=MOBILE_CELL_INDEX,
        company_cell_index=COMPANY_CELL_INDEX,
        address_cell_index=ADDRESS_CELL_INDEX,
        industry_cell_index=INDUSTRY_CELL_INDEX,
        sheet_name=DEFAULT_SHEET_NAME
):

    temp_file_name = 'list.xlsx'

    new_file = os.path.join(req.root_dir, temp_file_name)
    file.save(new_file)
    workbook = load_workbook(new_file, keep_vba=False, keep_links=False)
    sheet: Worksheet = workbook[sheet_name]
    print(f'max rows: {sheet.max_row}')

    # pick the tuple data
    data = [SentClient(
        name=v[name_cell_index].value,
        tel=f'+86{v[mobile_cell_index].value}' if add_86prefix else v[mobile_cell_index].value,
        address=v[address_cell_index].value if len(v[address_cell_index].value) > 7 else v[17].value,
        company=v[company_cell_index].value,
        industry=v[industry_cell_index].value,
    ) for v in sheet.rows]
    data.pop(0)
    for d in data:
        print(d)
    workbook.close()
    return data


def prepare_send_sms(sdk_id: str, template_id='881535', sign: str = '浩轩陈皮') -> models.SendSmsRequest:
    req = models.SendSmsRequest()
    req.SmsSdkAppid = sdk_id
    req.Sign = sign
    req.TemplateID = template_id
    return req


def waiting_follow_notify(seller_id: int, status: FollowStatusChoices, config):
    """
    :return: 返回seller关联的指定状态的需要提醒跟进的记录
    """
    days_ago = datetime.datetime.now() - datetime.timedelta(days=config[status])
    result = Following.query.filter_by(status=status).filter_by(follower_id=seller_id)
    return result.filter(Following.updated_at < days_ago)

def send_sms(numbers: List[str], req: models.SendSmsRequest, cred: credential.Credential):
    client = sms_client.SmsClient(cred, 'ap-guangzhou')
    req.PhoneNumberSet = numbers
    try:
        resp = client.SendSms(req)
        return resp.to_json_string()
    except TencentCloudSDKException as err:
        print(err)
        return False
