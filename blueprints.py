from flask import Blueprint, render_template, request, current_app
from openpyxl.worksheet.worksheet import Worksheet
from werkzeug.datastructures import FileStorage
from collections import namedtuple
from openpyxl import load_workbook
import json
import os


sms_app = Blueprint('sms', import_name='sms', url_prefix='/sms')
SentClient = namedtuple('SentClient', ['client_name', 'client_number'])


@sms_app.route('/test', methods=['GET', 'POST'])
def test_send():
    from app import prepare_send_sms, send_sms
    number = request.form.get('number', '+8613925114811')
    req = prepare_send_sms()
    return send_sms([number], req)


@sms_app.route('/add-template', methods=['GET', 'POST'])
def add_template():
    return 'Add Template View'


def parse_client_list(req, file: FileStorage):
    # const index for the file format
    NAME_CELL_INDEX = 0
    MOBILE_CELL_INDEX = 3
    COMPANY_CELL_INDEX = 4
    TEMP_FILE_NAME = 'list.xlsx'
    DEFAULT_SHEET_NAME = 'Sheet0'

    new_file = os.path.join(req.root_dir, TEMP_FILE_NAME)
    file.save(new_file)
    workbook = load_workbook(new_file, keep_vba=False, keep_links=False)
    sheet: Worksheet = workbook[DEFAULT_SHEET_NAME]
    print(f'max rows: {sheet.max_row}')

    # pick the tuple data
    data = [SentClient(
        client_name=v[COMPANY_CELL_INDEX].value,
        client_number=f'+86{v[MOBILE_CELL_INDEX].value}',
    ) for v in sheet.rows]
    data.pop(0)
    workbook.close()
    return data


@sms_app.route('/parse', methods=['GET', 'POST'])
def parse_to_csv():
    with_name = request.form.get('with-name', False)

    # TODO: complete this method, convert to csv file for download

    if with_name:
        pass
    else:
        pass
    return 'Not Implement'


@sms_app.route('/send', methods=['GET', 'POST'])
def send():
    if request.method == 'GET':
        return render_template('sms_send.html')

    try:
        from app import prepare_send_sms
        template_id = request.form.get('template_id')
        sign = request.form.get('sign')
        if template_id and sign:
            req = prepare_send_sms(template_id=template_id, sign=sign)
        else:
            req = prepare_send_sms()

        from app import send_sms
        data = parse_client_list(request, request.files['list'])
        for d in data:
            print(d)
        resp = json.loads(send_sms([v.client_number for v in data], req=req))
        return render_template('sms_finished.html', resp=resp)
    except Exception as ex:
        print(ex)
        return render_template('sms_send.html')


@sms_app.route('/send-single', methods=['GET', 'POST'])
def send_single():
    if request.method == 'GET':
        return render_template('sms_send_single.html')

    try:
        from app import prepare_send_sms
        number = '+86' + request.form['number']
        template_id = request.form.get('template_id')
        sign = request.form.get('sign')

        if template_id and sign:
            req = prepare_send_sms(template_id=template_id, sign=sign)
        else:
            req = prepare_send_sms()

        from app import send_sms
        resp = json.loads(send_sms([number], req=req))
        return render_template('sms_finished.html', resp=resp)
    except Exception as ex:
        print(ex)
        return render_template('sms_send_single.html')


@sms_app.route('/index', methods=['GET'])
def index():
    return render_template('sms_index.html')
